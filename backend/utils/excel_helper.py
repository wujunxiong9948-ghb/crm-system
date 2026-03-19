"""
Excel导入导出工具类
"""
import io
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class ExcelHelper:
    """Excel辅助类"""
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], headers: Dict[str, str], filename: str = None) -> io.BytesIO:
        """
        导出数据到Excel
        
        Args:
            data: 数据列表，每项是一个字典
            headers: 表头映射 {字段名: 显示名称}
            filename: 文件名（可选）
            
        Returns:
            BytesIO对象
        """
        wb = Workbook()
        ws = wb.active
        
        # 设置表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入表头
        for col_idx, (field, title) in enumerate(headers.items(), 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, item in enumerate(data, 2):
            for col_idx, field in enumerate(headers.keys(), 1):
                value = item.get(field, '')
                # 处理日期格式
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 自动调整列宽
        for col_idx, field in enumerate(headers.keys(), 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(headers[field])
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def import_from_excel(file_stream: io.BytesIO, headers: Dict[str, str]) -> List[Dict[str, Any]]:
        """
        从Excel导入数据
        
        Args:
            file_stream: 文件流
            headers: 表头映射 {字段名: 显示名称}
            
        Returns:
            数据列表
        """
        wb = load_workbook(file_stream)
        ws = wb.active
        
        # 获取表头
        header_row = []
        for cell in ws[1]:
            header_row.append(cell.value)
        
        # 创建字段映射（反向查找）
        field_map = {v: k for k, v in headers.items()}
        
        # 读取数据
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(header_row):
                    header = header_row[col_idx]
                    field = field_map.get(header)
                    if field:
                        item[field] = value
            if item:
                data.append(item)
        
        return data
    
    @staticmethod
    def create_template(headers: Dict[str, str]) -> io.BytesIO:
        """
        创建导入模板
        
        Args:
            headers: 表头映射 {字段名: 显示名称}
            
        Returns:
            BytesIO对象
        """
        wb = Workbook()
        ws = wb.active
        
        # 设置表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入表头
        for col_idx, (field, title) in enumerate(headers.items(), 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加示例数据行（灰色提示）
        example_font = Font(color='999999', italic=True)
        example_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        for col_idx, field in enumerate(headers.keys(), 1):
            cell = ws.cell(row=2, column=col_idx, value=f'示例{field}')
            cell.font = example_font
            cell.fill = example_fill
        
        # 添加说明sheet
        ws_info = wb.create_sheet('填写说明')
        ws_info['A1'] = '填写说明'
        ws_info['A1'].font = Font(bold=True, size=14)
        
        info_texts = [
            ('A3', '1. 请勿修改表头名称'),
            ('A4', '2. 日期格式：YYYY-MM-DD'),
            ('A5', '3. 必填字段必须填写'),
            ('A6', '4. 删除示例数据后再导入'),
            ('A8', '必填字段：'),
        ]
        
        for cell, text in info_texts:
            ws_info[cell] = text
        
        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20
        
        ws_info.column_dimensions['A'].width = 50
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

# 各模块的导出配置
EXPORT_CONFIG = {
    'customers': {
        'headers': {
            'id': 'ID',
            'name': '客户名称',
            'company': '公司名称',
            'phone': '电话',
            'email': '邮箱',
            'industry': '行业',
            'customer_type': '客户类型',
            'source': '来源',
            'status': '状态',
            'assigned_to': '负责人',
            'created_at': '创建时间',
        },
        'filename': '客户列表_{timestamp}.xlsx'
    },
    'opportunities': {
        'headers': {
            'id': 'ID',
            'name': '机会名称',
            'customer_name': '客户名称',
            'company': '公司',
            'project_type': '项目类型',
            'hotel_star': '酒店星级',
            'room_count': '客房数',
            'expected_value': '预计金额',
            'stage': '阶段',
            'probability': '概率%',
            'status': '状态',
            'assigned_to': '负责人',
            'expected_close_date': '预计成交日期',
            'created_at': '创建时间',
        },
        'filename': '销售机会_{timestamp}.xlsx'
    },
    'orders': {
        'headers': {
            'id': 'ID',
            'order_number': '订单编号',
            'customer_name': '客户名称',
            'opportunity_name': '关联机会',
            'total_amount': '总金额',
            'currency': '货币',
            'status': '状态',
            'payment_status': '支付状态',
            'order_date': '订单日期',
            'created_at': '创建时间',
        },
        'filename': '订单列表_{timestamp}.xlsx'
    },
    'products': {
        'headers': {
            'id': 'ID',
            'product_code': '产品编码',
            'description': '产品名称',
            'category': '分类',
            'material': '材质',
            'unit_price': '单价',
            'moq': '最小起订量',
            'specifications': '规格',
            'status': '状态',
        },
        'filename': '产品列表_{timestamp}.xlsx'
    }
}

# 导入模板配置
IMPORT_TEMPLATES = {
    'customers': {
        'headers': {
            'name': '客户名称 *',
            'company': '公司名称',
            'phone': '电话 *',
            'email': '邮箱',
            'industry': '行业',
            'customer_type': '客户类型',
            'source': '来源',
            'address': '地址',
        },
        'filename': '客户导入模板.xlsx'
    }
}
